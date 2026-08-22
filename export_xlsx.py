# -*- coding: utf-8 -*-
"""从 gauntlet_data.json 导出透视表格式 gauntlet_data.xlsx
- 9 sheets: 五区/四区 × 理论/高手/普通/自动 + 特殊跑法(明细)
- 行 = 大地图/小地图 (按大地图分组)
- 列 = 车辆 (高手档按星级拆列, 如 ssc★2/ssc★6; 无星级条目显示纯车名)
- 格 = 成绩(秒), 无数据留空; 普通/自动档填 ✓ 表示该车可用
- 特殊跑法 = 明细式 (大地图/小地图/车辆/星级/区-档/成绩/类型)
"""
import json, sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON = os.path.join(SCRIPT_DIR, 'gauntlet_data.json')
XLSX = os.path.join(SCRIPT_DIR, 'gauntlet_data.xlsx')

d = json.load(open(JSON, encoding='utf-8'))

HEADER_FILL = PatternFill('solid', fgColor='2F5496')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
ZONE_FILL = PatternFill('solid', fgColor='D9E2F3')   # 大地图分组行底色
TRACK_FONT = Font(size=10)
CAR_HEADER_FONT = Font(size=9, bold=True, color='1F3864')
CELL_FONT = Font(size=10)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')

def car_combo_key(c):
    """(车名, 星级) — 高手档拆星级; None 星显示纯车名"""
    return (c['name'], c.get('stars'))

# ---------- 列排序: 数据量优先, 其次平均速度, 无数据车按名字排右侧 ----------

def compute_car_stats(zone):
    """每车: (覆盖赛道数, 平均相对速度)
    数据源: 理论+高手档 非sc 有time条目, 同一赛道取该车最快成绩
    相对速度 = 赛道最快成绩 / 该车成绩 (1.0 = 该赛道最快, 越小越慢)
    """
    # 先算每条赛道在该区的最快非sc成绩
    track_fastest = {}
    for t in d['tracks']:
        best = None
        for tier in ('理论', '高手'):
            for e in t.get(zone, {}).get(tier, []):
                if e.get('sc') or e.get('time') is None:
                    continue
                if best is None or e['time'] < best:
                    best = e['time']
        if best is not None:
            track_fastest[(t['大地图'], t['小地图'])] = best
    # 每车每赛道最快成绩
    car_best = {}  # car -> {track_key: best_time}
    for t in d['tracks']:
        tk = (t['大地图'], t['小地图'])
        for tier in ('理论', '高手'):
            for e in t.get(zone, {}).get(tier, []):
                if e.get('sc') or e.get('time') is None:
                    continue
                tm = e['time']
                for c in e.get('cars', []):
                    n = c['name']
                    if tk not in car_best.setdefault(n, {}) or tm < car_best[n][tk]:
                        car_best[n][tk] = tm
    stats = {}
    for n, tracks in car_best.items():
        ratios = [track_fastest[tk] / tm for tk, tm in tracks.items() if tk in track_fastest]
        stats[n] = (len(tracks), sum(ratios) / len(ratios) if ratios else 0.0)
    return stats

def order_combos(combos, stats):
    """列序: 车名按 (数据量↓, 速度↓, 名字) 排, 同车不同星聚在一起按星级升序
    None 星(纯车名)放该车组末尾
    """
    names = {n for n, s in combos}
    def key(n):
        cnt, spd = stats.get(n, (0, 0.0))
        return (-cnt, -spd, n)
    ranked = sorted(names, key=key)
    cols = []
    for n in ranked:
        stars = sorted({s for nn, s in combos if nn == n}, key=lambda s: (s is None, s if s is not None else 0))
        for s in stars:
            cols.append((n, s))
    return cols

def build_pivot_sheet(ws, zone, tier, with_stars, show_check=False, stats=None, preset_cols=None):
    """透视表: 行=赛道, 列=车辆
    stats: compute_car_stats(zone) 结果, 用于数据量+速度排序
    preset_cols: 指定列序 (普通档镜像高手档)
    """
    # 收集列 (含 sc-only 车, 主表留空, 成绩见特殊跑法表)
    combos = set()
    for t in d['tracks']:
        for e in t.get(zone, {}).get(tier, []):
            for c in e.get('cars', []):
                combos.add(car_combo_key(c))
    if preset_cols is not None:
        # 普通档镜像高手档列序; 防御: 万一有差异, 多余列按名字追加尾部
        cols = [k for k in preset_cols if k in combos]
        extra = sorted([k for k in combos if k not in preset_cols], key=lambda k: (k[0], k[1] if k[1] is not None else 0))
        cols += extra
    else:
        cols = order_combos(combos, stats)

    # 表头
    ws.cell(1, 1, '大地图')
    ws.cell(1, 2, '小地图')
    for j, (name, stars) in enumerate(cols, 3):
        label = f'{name}★{stars}' if stars else name
        ws.cell(1, j, label)
    for j in range(1, len(cols) + 3):
        c = ws.cell(1, j)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    # 行: 按大地图分组, 组间空一行? 不, 直接连续+大地图合并单元格
    r = 2
    cur_map = None
    map_start = 2
    for t in d['tracks']:
        dm, xm = t['大地图'], t['小地图']
        if dm != cur_map:
            if cur_map is not None:
                ws.merge_cells(start_row=map_start, start_column=1, end_row=r - 1, end_column=1)
            cur_map = dm
            map_start = r
        # 构建该赛道单元格数据
        cell = {}
        for e in t.get(zone, {}).get(tier, []):
            if e.get('sc'):
                continue
            for c in e.get('cars', []):
                key = car_combo_key(c)
                val = e.get('time')
                if key in cell and val is not None and cell[key] is not None:
                    # 同格多值: 取更快(更小), 记录警告
                    cell[key] = min(cell[key], val)
                elif val is not None:
                    cell[key] = val
                elif key not in cell:
                    cell[key] = None
        ws.cell(r, 1, dm)
        ws.cell(r, 2, xm)
        for j, key in enumerate(cols, 3):
            v = cell.get(key)
            if show_check:
                # 普通/自动: 有条目填 ✓
                if key in cell:
                    ws.cell(r, j, '✓')
            else:
                if v is not None:
                    ws.cell(r, j, v)
        # 样式
        for j in range(1, len(cols) + 3):
            c = ws.cell(r, j)
            c.border = BORDER
            if j <= 2:
                c.font = TRACK_FONT
            else:
                c.font = CELL_FONT
                c.alignment = CENTER
        if dm == cur_map:
            ws.cell(r, 1).fill = ZONE_FILL
        r += 1
    if cur_map is not None:
        ws.merge_cells(start_row=map_start, start_column=1, end_row=r - 1, end_column=1)
        ws.cell(map_start, 1).fill = ZONE_FILL
        ws.cell(map_start, 1).alignment = CENTER

    # 列宽
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    for j in range(3, len(cols) + 3):
        ws.column_dimensions[get_column_letter(j)].width = 7.5
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols)+2)}{r-1}'
    return cols

def build_sc_sheet(ws):
    """特殊跑法 明细式"""
    header = ['大地图', '小地图', '车辆', '星级', '区-档', '成绩', '类型']
    for j, h in enumerate(header, 1):
        c = ws.cell(1, j, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    r = 2
    for t in d['tracks']:
        for zone in ['五区', '四区']:
            for tier in ['理论', '高手']:
                for e in t.get(zone, {}).get(tier, []):
                    if not e.get('sc'):
                        continue
                    for c in e.get('cars', []):
                        zt = f'{zone}{"理" if tier == "理论" else "高"}'
                        ws.cell(r, 1, t['大地图'])
                        ws.cell(r, 2, t['小地图'])
                        ws.cell(r, 3, c['name'])
                        ws.cell(r, 4, f'★{c["stars"]}' if c.get('stars') else '')
                        ws.cell(r, 5, zt)
                        ws.cell(r, 6, e.get('time'))
                        ws.cell(r, 7, e.get('sc_type') or 'sc')
                        for j in range(1, 8):
                            cc = ws.cell(r, j)
                            cc.border = BORDER
                            cc.font = CELL_FONT
                            if j in (4, 5, 6):
                                cc.alignment = CENTER
                        r += 1
    widths = [14, 14, 10, 8, 9, 9, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.auto_filter.ref = f'A1:G{r-1}'
    ws.freeze_panes = 'A2'
    return r - 2

wb = openpyxl.Workbook()
# 移除默认 sheet
default = wb.active
wb.remove(default)

stats = {}
for zone in ['五区', '四区']:
    car_stats = compute_car_stats(zone)
    high_cols = None
    for tier, show in [('理论', False), ('高手', False), ('普通', True), ('自动', True)]:
        ws = wb.create_sheet(f'{zone}_{tier}')
        preset = high_cols if tier == '普通' else None
        cols = build_pivot_sheet(ws, zone, tier, with_stars=(tier == '高手'), show_check=show, stats=car_stats, preset_cols=preset)
        if tier == '高手':
            high_cols = cols
        stats[f'{zone}_{tier}'] = len(cols)

ws = wb.create_sheet('特殊跑法')
sc_n = build_sc_sheet(ws)
stats['特殊跑法'] = sc_n

wb.save(XLSX)
print('saved:', XLSX)
for k, v in stats.items():
    print(f'  {k}: 列数/条目 = {v}')
