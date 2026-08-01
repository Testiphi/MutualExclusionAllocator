# -*- coding: utf-8 -*-
"""对比 gauntlet_data.xlsx(基准) 与 gauntlet_data.user.xlsx(用户版), 输出所有差异"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE = 'C:/Users/asus/.openclaw/workspace/tasks/mutex-alloc/repo/gauntlet_data.xlsx'
USER = 'C:/Users/asus/.openclaw/workspace/tasks/mutex-alloc/repo/gauntlet_data.user.xlsx'

wb_b = openpyxl.load_workbook(BASE)
wb_u = openpyxl.load_workbook(USER)

def norm(v):
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v

total = 0
for sn in wb_b.sheetnames:
    ws_b, ws_u = wb_b[sn], wb_u[sn]
    if sn not in wb_u.sheetnames:
        print(f'[{sn}] 用户版缺少此 sheet!')
        continue
    rows_b, rows_u = ws_b.max_row, ws_u.max_row
    cols_b, cols_u = ws_b.max_column, ws_u.max_column
    if (rows_b, cols_b) != (rows_u, cols_u):
        print(f'[{sn}] 尺寸不同: 基准 {rows_b}x{cols_b} vs 用户 {rows_u}x{cols_u}')
    # 逐格对比
    for r in range(1, max(rows_b, rows_u) + 1):
        for c in range(1, max(cols_b, cols_u) + 1):
            vb = norm(ws_b.cell(r, c).value)
            vu = norm(ws_u.cell(r, c).value)
            if vb != vu:
                total += 1
                # 行描述: 前两列是大地图/小地图(或表头)
                rlabel = ws_u.cell(r, 1).value if ws_u.cell(r, 1).value else f'row{r}'
                mlabel = ws_u.cell(r, 2).value if ws_u.cell(r, 2).value else ''
                clabel = ws_u.cell(1, c).value if ws_u.cell(1, c).value else f'col{c}'
                if total <= 200:
                    print(f'[{sn}] r{r} ({rlabel}/{mlabel}) 列[{clabel}]: 基准={vb!r} 用户={vu!r}')

print(f'\n共 {total} 处差异')
